import time as time
from pathlib import Path
from typing import Callable, Dict, List, Tuple

import astropy.constants as const
import astropy.units as u
import numpy as np
from astropy.modeling.models import BlackBody
from openpyxl import Workbook, load_workbook
from scipy.interpolate import interp1d
from scipy.special import j1

from .options import OPTIONS


def get_binning_windows(wavelength: np.ndarray) -> np.ndarray:
    """Gets all the binning windows."""
    skip_set = set()
    all_binning_windows = []
    for band in list(map(get_band, wavelength)):
        windows = getattr(OPTIONS.data.binning, band).value
        if band in skip_set:
            continue

        if isinstance(windows, (list, tuple, np.ndarray)):
            all_binning_windows.extend(windows)
            skip_set.add(band)
        else:
            all_binning_windows.append(windows)
    return all_binning_windows * u.um


def create_adaptive_bins(wavelength_range_full: List[float], wavelength_range: List[float],
                         bin_window_in: float, bin_window_out: float) -> np.ndarray:
    """Create an adaptive binning wavelength grid.

    Parameters
    ----------
    wavelength_range_full : list of float
    wavelength_range : list of float
    bin_window_in : float
    bin_window_out : float

    Returns
    -------
    bin_centres : numpy.ndarray
    """
    range_full_min, range_full_max = wavelength_range_full
    range_min, range_max = wavelength_range

    if range_full_min >= range_full_max or range_min >= range_max:
        raise ValueError("Invalid wavelength ranges.")
    if not (range_full_min <= range_min and range_max <= range_full_max):
        raise ValueError("Wavelength range must be within the full wavelength range.")

    below_range = np.arange(range_full_min, range_min - bin_window_out / 2, bin_window_out)
    within_range = np.arange(range_min, range_max, bin_window_in)
    above_range = np.arange(range_max + bin_window_out / 2, range_full_max)
    all_edges = np.unique(np.concatenate((below_range, within_range, above_range)))
    windows = np.concatenate((np.full(below_range.shape, bin_window_out),
                              np.full(within_range.shape, bin_window_in),
                              np.full(above_range.shape, bin_window_out)))
    return all_edges, windows


def compare_angles(angle1: u.Quantity, angle2: u.Quantity) -> complex:
    """Subtracts two angles and makes sure the are between -np.pi and +np.pi."""
    if isinstance(angle1, u.Quantity):
        angle1 = angle1.to(u.rad).value

    if isinstance(angle2, u.Quantity):
        angle2 = angle2.to(u.rad).value

    diff = angle1 - angle2
    diff[diff > np.pi] -= 2 * np.pi
    diff[diff < -np.pi] += 2 * np.pi
    return diff


def windowed_linspace(start: float, end: float, window: float) -> np.ndarray:
    """Creates a numpy.linspace with a number of points so that the windowing doesn't overlap"""
    return np.linspace(start, end, int((end - start) // (2 * window)) + 1)


def get_band_limits(band: str) -> Tuple[float, float]:
    """Gets the limits of the respective band"""
    match band:
        case "hband":
            return 1.5, 1.8
        case "kband":
            return 1.9, 2.5
        case "lband":
            return 2.8, 3.99
        case "mband":
            return 4.0, 6.0
        case "nband":
            return 7.5, 16.0
    return 0, 0


def get_band(wavelength: u.um) -> str:
    """Gets the band of the (.fits)-file."""
    wavelength = wavelength.value if isinstance(wavelength, u.Quantity) else wavelength
    wl_min, wl_max = wavelength.min(), wavelength.max()
    if wl_min > 1.5 and wl_max < 1.8:
        return "hband"
    if wl_min > 1.9 and wl_max < 2.5:
        return "kband"
    if wl_min > 2.8 and wl_max < 4.0:
        return "lband"
    if wl_min >= 4.0 and wl_max < 6.0:
        return "mband"
    if wl_min > 2.8 and wl_max < 6:
        return "lmband"
    if wl_min > 7.5 and wl_max < 16.0:
        return "nband"
    return "unknown"


def smooth_interpolation(
    interpolation_points: np.ndarray,
    grid: np.ndarray,
    values: np.ndarray,
    kind: str | None = None,
    fill_value: str | None = None,
) -> np.ndarray:
    """Rebins the grid to a higher factor and then interpolates and averages
    to the original grid.

    Parameters
    ----------
    interpolation_points : numpy.ndarray
        The points to interpolate to.
    points : numpy.ndarray
        The points to interpolate from.
    values : numpy.ndarray
        The values to interpolate.
    """
    kind = OPTIONS.data.interpolation.kind if kind is None else kind
    fill_value = (
        OPTIONS.data.interpolation.fill_value if fill_value is None else fill_value
    )
    points = interpolation_points.flatten()
    windows = get_binning_windows(points).value
    interpolation_grid = (
        np.linspace(-1, 1, OPTIONS.data.interpolation.dim) * windows[:, np.newaxis] / 2
    ).T + points
    return (
        np.interp(interpolation_grid, grid, values)
        .mean(axis=0)
        .reshape(interpolation_points.shape)
    )


def take_time_average(
    func: Callable, *args, nsteps: int = 10
) -> Tuple[Callable, float]:
    """Takes a time average of the code."""
    execution_times = []
    for _ in range(nsteps):
        time_st = time.perf_counter()
        return_val = func(*args)
        execution_times.append(time.perf_counter() - time_st)
    return return_val, np.array(execution_times).mean()


def get_indices(
    values: np.ndarray, array: np.ndarray, windows: np.ndarray | float | None = None
) -> List[np.ndarray]:
    """Gets the indices of values occurring in a numpy array
    and returns it in a list corresponding to the input values.

    Parameters
    ----------
    values : float
        The values to find.
    array : numpy.ndarray
        The array to search in.
    window : float
        The window around the value to search in.
    """
    array = array.value if isinstance(array, u.Quantity) else array
    values = values.value if isinstance(values, u.Quantity) else values
    windows = windows.value if isinstance(windows, u.Quantity) else windows
    if not isinstance(values, (list, tuple, np.ndarray)):
        values = [values]

    if windows is not None:
        if isinstance(windows, (list, tuple, np.ndarray)):
            indices = [np.where(((v - w / 2) < array) & ((v + w / 2) > array))[0] for v, w in zip(values, windows)]
        else:
            indices = [np.where(((v - windows / 2) < array) & ((v + windows / 2) > array))[0] for v in values]
    else:
        indices = []
        for value in values:
            index = np.where(array == value)[0]
            if index.size == 0:
                if value < array[0] or value > array[-1]:
                    indices.append(index.astype(int).flatten())
                    continue

                index = np.where(array == min(array, key=lambda x: abs(x - value)))[0]

            indices.append(index.astype(int).flatten())
    return indices


def compute_photometric_slope(wavelengths: u.um, temperature: u.K) -> np.ndarray:
    """Computes the photometric slope of the data from
    the effective temperature of the star.

    Parameters
    ----------
    wavelengths : astropy.units.um
        The wavelengths of the data.
    temperature : astropy.units.K
        The effective temperature of the star.

    Returns
    -------
    photometric_slope : numpy.ndarray
    """
    temperature = u.Quantity(temperature, u.K)
    wavelengths = u.Quantity(wavelengths, u.um)
    nu = (const.c / wavelengths.to(u.m)).to(u.Hz)
    blackbody = BlackBody(temperature)
    return np.gradient(np.log(blackbody(nu).value), np.log(nu.value))


def compute_stellar_radius(luminosity: u.Lsun, temperature: u.K) -> u.Rsun:
    """Calculates the stellar radius from the luminosity and temperature."""
    luminosity, temperature = (
        u.Quantity(luminosity, u.Lsun),
        u.Quantity(temperature, u.K),
    )
    return np.sqrt(
        luminosity.to(u.W) / (4 * np.pi * const.sigma_sb * temperature**4)
    ).to(u.Rsun)


def angular_to_distance(angular_diameter: u.mas, distance: u.pc) -> u.m:
    """Converts an angular diameter of an object at a certain distance
    from the observer from mas to meters.

    Parameters
    ----------
    angular_diameter : astropy.units.mas
        The angular diameter of an object.
    distance : astropy.units.pc
        The distance to the object.

    Returns
    -------
    diameter : astropy.units.m
        The diameter of the object.

    Notes
    -----
    The formula for the angular diameter small angle approximation is

    .. math:: d = \\delta*D

    where 'd' is the diameter of the object and 'D' is the distance from the
    observer to the object and ..math::`\\delta` is the angular diameter.
    """
    return angular_diameter.to(u.rad).value * distance.to(u.m)


def distance_to_angular(diameter: u.au, distance: u.pc) -> u.mas:
    """Converts an angular diameter of an object at a certain distance
    from the observer from mas to meters.

    Parameters
    ----------
    diameter : astropy.units.au
        The diameter of an object.
    distance : astropy.units.pc
        The distance to the object.

    Returns
    -------
    diameter : astropy.units.mas
        The diameter of the object.

    Notes
    -----
    The formula for the angular diameter small angle approximation is

    .. math:: \\delta = \\frac{d}{D}

    where 'd' is the diameter of the object and 'D' is the distance from the
    observer to the object and ..math::`\\delta` is the angular diameter.
    """
    return ((diameter.to(u.m) / distance.to(u.m)) * u.rad).to(u.mas)


def compute_effective_baselines(
    ucoord: u.m,
    vcoord: u.m,
    inclination: u.Quantity[u.one] | None = None,
    pos_angle: u.Quantity[u.deg] | None = None,
    longest: bool | None = False,
    return_zero: bool | None = True,
) -> Tuple[u.Quantity[u.m], u.Quantity[u.one]]:
    """Calculates the effective baselines from the projected baselines
    in mega lambda.

    Parameters
    ----------
    ucoord: astropy.units.m
        The u coordinate.
    vcoord: astropy.units.m
        The v coordinate.
    inclination: astropy.units.one
        The inclinatin induced compression of the x-axis.
    pos_angle: astropy.units.deg
        The positional angle of the object
    longest : bool, optional
        If True, the longest baselines are returned.
    return_zero : bool, optional
        If True, the zero-frequency is returned as well.

    Returns
    -------
    baselines : astropy.units.m
        Returns the effective baselines.
    baselines_angles : astropy.units.rad
        Returns the effective baseline angles.
    """
    ucoord, vcoord = map(lambda x: u.Quantity(x, u.m), [ucoord, vcoord])
    if pos_angle is not None:
        pos_angle = u.Quantity(pos_angle, u.deg).to(u.rad)
        inclination = u.Quantity(inclination, u.one)

        ucoord_eff = ucoord * np.cos(pos_angle) - vcoord * np.sin(pos_angle)
        vcoord_eff = ucoord * np.sin(pos_angle) + vcoord * np.cos(pos_angle)
    else:
        ucoord_eff, vcoord_eff = ucoord, vcoord

    if inclination is not None:
        ucoord_eff *= inclination

    baselines_eff = np.hypot(ucoord_eff, vcoord_eff)
    baseline_angles_eff = np.arctan2(vcoord_eff, ucoord_eff)

    if longest:
        indices = baselines_eff.argmax(0)
        iteration = np.arange(baselines_eff.shape[1])
        baselines_eff = baselines_eff[indices, iteration]
        baseline_angles_eff = baseline_angles_eff[indices, iteration]

    if not return_zero:
        if len(baselines_eff.shape) > 1:
            baselines_eff = baselines_eff[:, 1:]
            baseline_angles_eff = baseline_angles_eff[:, 1:]
        else:
            baselines_eff = baselines_eff[1:]
            baseline_angles_eff = baseline_angles_eff[1:]

    return baselines_eff.squeeze(), baseline_angles_eff.squeeze()


def binary(
    dim: int,
    pixel_size: u.mas,
    flux1: u.Jy,
    flux2: u.Jy,
    position1: u.mas,
    position2: u.mas,
) -> u.Jy:
    """The image of a binary.

    Parameters
    ----------
    dim : float
        The image's dimension (px).
    pixel_size : astropy.units.mas
        The size of a pixel in the image.
    flux1 : astropy.units.Jy
        The main component's flux.
    flux2 : astropy.units.Jy
        The companion's flux.
    position1 : astropy.units.m
        The main component's (x, y)-coordinates.
    position2 : astropy.units.m
        The companion's (x, y)-coordinates.
    wavelength : astropy.units.um

    Returns
    -------
    image : astropy.units.Jy
    """
    v = np.linspace(-0.5, 0.5, dim, endpoint=False) * pixel_size.to(u.mas) * dim
    image = np.zeros((dim, dim)) * u.Jy
    position1 = (
        np.array(list(map(lambda x: np.where(v == x), position1))).flatten().tolist()
    )
    position2 = (
        np.array(list(map(lambda x: np.where(v == x), position2))).flatten().tolist()
    )
    image[position1[1]][position1[0]] = flux1
    image[position2[1]][position2[0]] = flux2
    return image


def binary_vis(
    flux1: u.Jy,
    flux2: u.Jy,
    ucoord: u.m,
    vcoord: u.m,
    position1: u.mas,
    position2: u.mas,
    wavelength: u.um,
) -> np.ndarray:
    """The complex visibility function of a binary star.

    Parameters
    ----------
    flux1 : astropy.units.Jy
        The main component's flux.
    flux2 : astropy.units.Jy
        The companion's flux.
    position1 : astropy.units.m
        The main component's (x, y)-coordinates.
    position2 : astropy.units.m
        The companion's (x, y)-coordinates.
    wavelength : astropy.units.um

    Returns
    -------
    complex_visibility_function : numpy.ndarray
    """
    ucoord, vcoord = map(lambda x: x / wavelength.to(u.m), [ucoord, vcoord])
    xy_and_uv = list(
        map(
            lambda x: ucoord * x.to(u.rad)[0] + vcoord * x.to(u.rad)[1],
            [position1, position2],
        )
    )
    return flux1.value * np.exp(
        2 * np.pi * 1j * xy_and_uv[0].value
    ) + flux2.value * np.exp(2 * np.pi * 1j * xy_and_uv[1].value)


def uniform_disk(
    pixel_size: u.mas, dim: int, diameter: u.Quantity[u.mas] | None = None
) -> u.one:
    """The brightness profile of a uniform disk.

    Parameters
    ----------
    pixel_size : astropy.units.mas
        The size of a pixel in the image.
    dim : float
        The image's dimension [px].
    diameter : astropy.units.mas, optional
        The uniform disk's diameter.

    Returns
    -------
    radial_profile : astropy.units.one
    """
    if diameter is not None:
        v = np.linspace(-0.5, 0.5, dim, endpoint=False) * pixel_size.to(u.mas) * dim
        x_arr, y_arr = np.meshgrid(v, v)
        radius = np.hypot(x_arr, y_arr) < diameter / 2
    else:
        radius = np.ones((dim, dim)).astype(bool)
        diameter = 1 * u.mas
    return 4 * u.one * radius / (np.pi * diameter.value**2)


def uniform_disk_vis(
    diameter: u.mas, ucoord: u.m, vcoord: u.m, wavelength: u.um
) -> np.ndarray:
    """Defines the complex visibility function of a uniform disk.

    Parameters
    ----------
    diameter : astropy.units.mas
        The uniform disk's diameter.
    ucoord : astropy.units.m
        The u-coordinates.
    vcoord : astropy.units.m
        The v-coordinates.
    wavelength : astropy.units.um
        The wavelength for the spatial frequencies' unit conversion.

    Returns
    -------
    complex_visibility_function : numpy.ndarray
    """
    rho = np.hypot(ucoord, vcoord) / wavelength.to(u.m)
    return (
        2
        * j1(np.pi * rho * diameter.to(u.rad).value)
        / (np.pi * diameter.to(u.rad).value * rho)
    )


def make_workbook(file: Path, sheets: Dict[str, List[str]]) -> None:
    """Creates an (.xslx)-sheet with subsheets.

    Parameters
    ----------
    file : Path
    sheets : dict of list
        A dictionary having the sheet name as they key
        and a list of columns as the value.
    """
    file_existed = True
    if file.exists():
        workbook = load_workbook(file)
    else:
        workbook = Workbook()
        file_existed = False
    for index, (sheet, columns) in enumerate(sheets.items()):
        if sheet not in workbook.sheetnames:
            if index == 0 and not file_existed:
                worksheet = workbook.active
                worksheet.title = sheet
            else:
                worksheet = workbook.create_sheet(title=sheet)
        else:
            worksheet = workbook[sheet]
        worksheet.delete_rows(1, worksheet.max_row)
        for col_idx, column_name in enumerate(columns, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = column_name
    workbook.save(file)


def qval_to_opacity(qval_file: Path) -> u.cm**2 / u.g:
    """Reads a qval file, then calculates and returns the
    opacity.

    Parameters
    ----------
    qval_file : pathlib.Path

    Returns
    -------
    opacity : astropy.units.cm**2/u.g

    Notes
    -----
    The qval-files give the grain size in microns and the
    density in g/cm^3.
    """
    with open(qval_file, "r+", encoding="utf8") as file:
        _, grain_size, density = map(float, file.readline().strip().split())
    wavelength_grid, qval = np.loadtxt(
        qval_file, skiprows=1, unpack=True, usecols=(0, 1)
    )
    return wavelength_grid * u.um, 3 * qval / (
        4 * (grain_size * u.um).to(u.cm) * (density * u.g / u.cm**3)
    )


def restrict_wavelength(
    wavelength: np.ndarray, array: np.ndarray, wavelength_range: u.um
) -> Tuple[np.ndarray, np.ndarray]:
    """Restricts the wavelength for an input range."""
    indices = (wavelength > wavelength_range[0]) & (wavelength < wavelength_range[1])
    return wavelength[indices], array[indices]


def restrict_phase(phase: np.ndarray) -> np.ndarray:
    """Restricts the phase to [-180, 180] degrees."""
    restricted_phase = phase % 360
    restricted_phase[restricted_phase > 180] -= 360
    return restricted_phase


def get_opacity(
    source_dir: Path,
    weights: np.ndarray,
    names: List[str],
    method: str,
    individual: bool = False,
    **kwargs,
) -> Tuple[np.ndarray, np.ndarray]:
    """Gets the opacity from input parameters."""
    grf_dict = {
        "olivine": "Olivine",
        "pyroxene": "MgPyroxene",
        "forsterite": "Forsterite",
        "enstatite": "Enstatite",
        "silica": "Silica",
    }

    files = []
    for name in names:
        name = name.lower()
        for size in ["small", "large"]:
            if method == "grf":
                size = 0.1 if size == "small" else 2
                file_name = f"{grf_dict[name]}{size:.1f}.Combined.Kappa"
            else:
                size = "Big" if size == "large" else size.title()
                file_name = f"{size}{name.title()}.kappa"

            files.append(source_dir / method / file_name)

    usecols = (0, 2) if method == "grf" else (0, 1)
    wl, opacity = load_data(files, usecols=usecols, **kwargs)

    if individual:
        return wl, opacity

    opacity = (opacity * weights[:, np.newaxis]).sum(axis=0)
    return wl, opacity


def load_data(
    files: Path | List[Path],
    load_func: Callable | None = None,
    comments: str = "#",
    skiprows: int = 1,
    usecols: Tuple[int, int] = (0, 1),
    method: str = "shortest",
    kind: str | None = None,
    fill_value: str | None = None,
) -> Tuple[np.ndarray, np.ndarray]:
    """Loads data from a file.

    Can either be one or multiple files, but in case
    of multiple files they need to have the same structure
    and size (as they will be converted to numpy.ndarrays).

    Parameters
    ----------
    files : list of pathlib.Path
        The files to load the data from.
    load_func : callable, optional
        The function to load the data with.
    comments : str, optional
        Comment identifier.
    skiprows : str, optional
        The rows to skip.
    usecols : tuple of int, optional
        The columns to use.
    method : str, optional
        The grid to interpolate/extrapolate the data on.
        Default is 'shortest'. Other option is 'longest' or
        'median'.
    kind : str, optional
        The interpolation kind.
        Default is 'cubic'.
    fill_value : str, optional
        If "extrapolate", the data is extrapolated.
        Default is None.

    Returns
    -------
    wavelength_grid : numpy.ndarray
    data : numpy.ndarray
    """
    kind = OPTIONS.data.interpolation.kind if kind is None else kind
    fill_value = (
        OPTIONS.data.interpolation.fill_value if fill_value is None else fill_value
    )

    files = files if isinstance(files, list) else [files]
    wavelength_grids, contents = [], []
    for file in files:
        if load_func is not None:
            wavelengths, content = load_func(file)
        else:
            wavelengths, content = np.loadtxt(
                file, skiprows=skiprows, usecols=usecols, comments=comments, unpack=True
            )

        if isinstance(wavelengths, u.Quantity):
            wavelengths = wavelengths.value
            content = content.value

        wavelength_grids.append(wavelengths)
        contents.append(content)

    sizes = [np.size(wl) for wl in wavelength_grids]
    if method == "longest":
        wavelength_grid = wavelength_grids[np.argmax(sizes)]
    elif method == "shortest":
        wavelength_grid = wavelength_grids[np.argmin(sizes)]
    else:
        wavelength_grid = wavelength_grids[
            np.median(sizes).astype(int) == wavelength_grids
        ]

    data = []
    for wavelengths, content in zip(wavelength_grids, contents):
        if np.array_equal(wavelengths, wavelength_grid):
            data.append(content)
            continue

        data.append(
            interp1d(wavelengths, content, kind=kind, fill_value=fill_value)(
                wavelength_grid
            )
        )

    return wavelength_grid.squeeze(), np.array(data).squeeze()


def linearly_combine_data(data: np.ndarray, weights: u.one) -> np.ndarray:
    """Linearly combines multiple opacities by their weights.

    Parameters
    ----------
    data : numpy.ndarray
        The data to be linearly combined.
    weights : u.one
        The weights for the different data.

    Returns
    -------
    numpy.ndarray
    """
    return np.sum(data * weights[:, np.newaxis], axis=0)


# TODO: Remove the check for the ucoord here
def broadcast_baselines(
    wavelength: u.um, baselines: u.m, baseline_angles: u.rad, ucoord: u.m
) -> Tuple[u.Quantity[u.um], u.Quantity[1 / u.rad], u.Quantity[u.rad]]:
    """Broadcasts the baselines to the correct shape depending on
    the input ucoord shape."""
    wavelength = wavelength[:, np.newaxis]
    if ucoord.shape[0] == 1:
        baselines = (baselines / wavelength.to(u.m)).value
        baselines = baselines[..., np.newaxis]
        baseline_angles = baseline_angles[np.newaxis, :, np.newaxis]
    else:
        wavelength = wavelength[..., np.newaxis]
        baselines = (baselines[np.newaxis, ...] / wavelength.to(u.m)).value
        baselines = baselines[..., np.newaxis]
        baseline_angles = baseline_angles[np.newaxis, ..., np.newaxis]
    baseline_angles = u.Quantity(baseline_angles, unit=u.rad)
    return wavelength, baselines / u.rad, baseline_angles


def compute_vis(vis: np.ndarray) -> np.ndarray:
    """Computes the visibilities from the visibility function."""
    return np.abs(vis).astype(OPTIONS.data.dtype.real)


def compute_t3(vis: np.ndarray) -> np.ndarray:
    """Computes the closure phase from the visibility function."""
    if vis.size == 0:
        return np.array([])

    vis /= vis[:, :, 0][..., np.newaxis].real
    bispectrum = vis[:, 0] * vis[:, 1] * vis[:, 2].conj()
    return np.angle(bispectrum, deg=True).real
